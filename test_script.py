import csv
import zipfile
from pypdf import PdfReader
from conftest import ZIPPED_RESOURCES
from openpyxl import load_workbook


def test_read_pdf():
    with zipfile.ZipFile(ZIPPED_RESOURCES) as zip_file:
        with zip_file.open('file.pdf') as file:

            reader = PdfReader(file)
            text = reader.pages[0].extract_text()

            assert 'Здравствуйте!' in text


def test_read_xlsx():
    with zipfile.ZipFile(ZIPPED_RESOURCES) as zip_file:
        with zip_file.open('file.xlsx') as file:

            workbook = load_workbook(file)
            sheet = workbook.active
            expected_text = sheet.cell(row=2, column=6).value

            assert expected_text == 'Облигация'


def test_read_csv():
    with zipfile.ZipFile(ZIPPED_RESOURCES) as zip_file:
        with zip_file.open('file.csv') as file:

            content = file.read().decode('utf-8-sig')
            csvreader = list(csv.reader(content.splitlines()))
            second_row = csvreader[1]

            assert 'admin@zaleycash.com' in second_row[4]


